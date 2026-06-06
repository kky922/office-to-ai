"""Excel reader (.xlsx/.xlsm) using openpyxl in read_only/stream mode."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .common import Block, ImageAsset


def _escape(v) -> str:
    if v is None:
        return ""
    return str(v).replace("|", "\\|").replace("\r\n", " ").replace("\n", " ")


def _row_to_md(headers: list[str], row: list) -> str:
    return "| " + " | ".join(_escape(c) for c in (row + [""] * (len(headers) - len(row)))[: len(headers)]) + " |"


def _header_md(headers: list[str]) -> str:
    head = "| " + " | ".join(_escape(h) for h in headers) + " |"
    sep = "|" + "|".join(" --- " for _ in headers) + "|"
    return head + "\n" + sep


def read(path: str | Path, rows_per_section: int = 200) -> tuple[list[Block], list[ImageAsset]]:
    path = Path(path)
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    blocks: list[Block] = [
        Block(
            kind="heading",
            md=f"# {path.name}\n\n> Excel 문서 · 시트 {len(wb.sheetnames)}개\n",
            source=path.name,
        )
    ]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter: Iterable[tuple] = ws.iter_rows(values_only=True)
        try:
            headers = [_escape(c) if c is not None else "" for c in next(rows_iter)]
        except StopIteration:
            blocks.append(
                Block(
                    kind="sheet-section",
                    md=f"## 시트: {sheet_name}\n\n_(빈 시트)_\n",
                    source=f"시트: {sheet_name}",
                )
            )
            continue

        if not headers:
            headers = [""]

        blocks.append(Block(kind="heading", md=f"## 시트: {sheet_name}", source=f"시트: {sheet_name}"))

        buf: list[list] = []
        first_row_no = 2
        row_no = first_row_no

        def _flush() -> None:
            nonlocal buf, first_row_no
            if not buf:
                return
            label = f"시트: {sheet_name} · 행 {first_row_no}-{first_row_no + len(buf) - 1}"
            body = "\n".join(_row_to_md(headers, list(r)) for r in buf)
            md = f"### {label}\n\n{_header_md(headers)}\n{body}\n"
            blocks.append(Block(kind="sheet-section", md=md, source=label))
            first_row_no = first_row_no + len(buf)
            buf = []

        for row in rows_iter:
            buf.append(list(row))
            row_no += 1
            if len(buf) >= rows_per_section:
                _flush()
        _flush()
        if first_row_no == 2 and not buf:
            # only header was present
            blocks.append(
                Block(
                    kind="sheet-section",
                    md=f"{_header_md(headers)}\n_(데이터 행 없음)_\n",
                    source=f"시트: {sheet_name} · 헤더만",
                )
            )
    wb.close()
    return blocks, []
